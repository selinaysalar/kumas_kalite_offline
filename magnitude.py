import os
import cv2
import numpy as np
from openpyxl import *
from matplotlib import pyplot as plt
from os import listdir
from os.path import isfile, join
import xlsxwriter
import numpy as np
from skimage.feature import greycomatrix, greycoprops
import numpy as np
import matplotlib.pylab as plt
import pandas as pd

def fourier(img):

    f = np.fft.fft2(img)
    fshift = np.fft.fftshift(f)
    amplitude=np.abs(fshift)
    spectrum =  np.log(1+amplitude)
    return spectrum

def fft_transform (image_list):

    fft_data = []
    fft_freq = []
    power_spec = []

    for image in image_list:
        image=preprocessing2(image)[3]
        fft_window = np.fft.fft(image)
        fft_data.append(fft_window)

        freq  = np.fft.fftfreq(np.array(image).shape[-1], d=0.01)
        fft_freq.append(freq )

        fft_ps = np.abs(fft_window)**2
        power_spec.append(fft_ps)

    return  fft_ps,power_spec




def preprocessing2(img):
    gray2=cv2.cvtColor(img,cv2.COLOR_RGB2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    cl1 = clahe.apply(gray2)
    median2=cv2.medianBlur(cl1,5)
    kernel = np.ones((15, 15), np.uint8)
    morpho = cv2.dilate(median2, kernel, iterations=1)

    return gray2,cl1,median2,morpho

def preprocessing(img):
    img=cv2.cvtColor(img,cv2.COLOR_RGB2GRAY)
    equal = cv2.equalizeHist(img)
    ret,th=cv2.threshold(equal,253,0,cv2.THRESH_TOZERO)

    gamapowder=cv2.pow(th,80)
    ret,th1= cv2.threshold(gamapowder,251,255,cv2.THRESH_BINARY_INV)
    median=cv2.medianBlur(th1,3)

    return img,equal,th,gamapowder,th1,median

def viewGlcmPreprocessing(img):
    g1,g2,g3,g4=preprocessing2(img)
    plt.figure('Gray-Level Co-occuorence Matris')
    plt.subplot(2, 2, 1)
    plt.imshow(g1, cmap='gray')
    plt.title('Gray-scale image', size=5)

    plt.subplot(2, 2, 2)
    plt.imshow(g2, cmap='gray')
    plt.title('Clache', size=5)


    plt.subplot(2, 2, 3)
    plt.imshow(g4, cmap='gray')
    plt.title('Median', size=5)

    plt.subplot(2, 2, 4)
    plt.imshow(g4, cmap='gray')
    plt.title('Morphological', size=5)
    plt.show()

def template_match(img,template):

    w, h = template.shape[::-1]
    result = cv2.matchTemplate(img, template, cv2.TM_CCOEFF_NORMED)
    location = np.where(result >= 0.65)

    for point in zip(*location[::-1]):
      cv2.rectangle(img, point, (point[0] + w, point[1] + h), (152, 255, 0), 3)

    return img

def viewGraphFourier(img):

    t1,t2,t3=viewTemp(img)
    p1,p2,p3,p4,p5,p6=preprocessing(img)
    f1=fourier(p6)
    X, Y = f1.shape


    plt.figure('Image processing')
    plt.subplot(3, 3, 1)
    plt.imshow(img, cmap='gray')
    plt.title('Orijinal fabric image', size=5)

    plt.subplot(3, 3, 2)
    plt.imshow(p1, cmap='gray')
    plt.title('Gray-scale image', size=5)

    plt.subplot(3, 3, 3)
    plt.imshow(p2, cmap='gray')
    plt.title('Histogram-equalized  image',size=5)

    plt.subplot(3, 3, 4)
    plt.imshow(p3, cmap='gray')
    plt.title("Threshold1", size=5)

    plt.subplot(3, 3, 5)
    plt.imshow(p4, cmap='gray')
    plt.title('GamaPowder', size=5)

    plt.subplot(3, 3, 6)
    plt.imshow(p5, cmap='gray')
    plt.title('Threshold2',size=5)

    plt.subplot(3, 3, 7)
    plt.imshow(t1, cmap='gray')
    plt.title('Hole-defect', size=5)

    plt.subplot(3, 3, 8)
    plt.imshow(t2, cmap='gray')
    plt.title("Vertical defect", size=5)

    plt.subplot(3, 3, 9)
    plt.imshow(t3,cmap='gray')
    plt.title("Horizontal defect", size=5)

    plt.show()

def viewTemp(img):

    img_gr=cv2.cvtColor(img,cv2.COLOR_RGB2GRAY)
    img_hole =cv2.cvtColor(img,cv2.COLOR_RGB2GRAY)
    img_vertical=cv2.cvtColor(img,cv2.COLOR_RGB2GRAY)
    img_horizontal = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)

    template_hole = cv2.imread('template/defect_hole_64.jpg', cv2.IMREAD_GRAYSCALE)
    template_vertical = cv2.imread('template/template_vertical.jpg', cv2.IMREAD_GRAYSCALE)
    template_horizontal = cv2.imread('template/ho1.jpg', cv2.IMREAD_GRAYSCALE)

    t1 = template_match(img_hole, template_hole)
    t2 = template_match(img_vertical, template_vertical)
    t3 = template_match(img_horizontal, template_horizontal)


    plt.figure("Template-matching Algorithm")
    plt.subplot(1, 4, 1)
    plt.imshow(img, cmap='gray')
    plt.title("Fabric")

    plt.subplot(1, 4, 2)
    plt.imshow(t1,cmap='gray')
    plt.title("Hole-Defect")

    plt.subplot(1, 4, 3)
    plt.imshow(t2, cmap='gray')
    plt.title("Vertical-Defect")

    plt.subplot(1, 4, 4)
    plt.imshow(t3, cmap='gray')
    plt.title("Horizontal-Defect")
    plt.show()
    return t1,t2,t3
def zig_zag(array):

    r, c = array.shape
    i = j = 0
    array2 = []
    array2.append(array[i, j])
    while (i != r - 1 or j != c - 1):
        j += 1
        array2.append(array[i, j])
        while (j != 0):
            i += 1
            j -= 1
            array2.append(array[i, j])
        if (i != r - 1 or j != 0):
            i += 1
            array2.append(array[i, j])
            while (i != 0):
                j += 1
                i -= 1
                array2.append(array[i, j])
        else:
            j += 1
            array2.append(array[i, j])
            while (i != r - 1 or j != c - 1):
                while (j != c - 1):
                    i -= 1
                    j += 1
                    array2.append(array[i, j])
                i += 1
                array2.append(array[i, j])
                while (i != r - 1):
                    j -= 1
                    i += 1
                    array2.append(array[i, j])
                j += 1
                array2.append(array[i, j])
    return array2

def read_excel(filename):

    read_file = pd.read_excel(filename)
    array = np.asarray(read_file)

    return array
def imageprocessing(image_list):

    images=[]
    for ite in range(len(image_list)):
        img = image_list[ite]
        img=cv2.cvtColor(img,cv2.COLOR_RGB2GRAY)
        equal = cv2.equalizeHist(img)
        ret,th=cv2.threshold(equal,253,0,cv2.THRESH_TOZERO)
        gamapowder=cv2.pow(th,80)
        ret,th1= cv2.threshold(gamapowder,251,255,cv2.THRESH_BINARY_INV)
        median=cv2.medianBlur(th1,3)
        if img is not None:
            images.append(median)
    return images

def Glcm(list_of_images,filename):
    workbook = xlsxwriter.Workbook(filename)
    sheet = workbook.add_worksheet('DCT')


    if sheet == '\0':
        print('error in file creation')
    else:
        print('GLCM file created')
        for ite in  range(len(list_of_images)):
            img = list_of_images[ite]
            img = preprocessing(img)[5]
            bins = np.array([0, 16, 32, 48, 64, 80, 96, 112, 128, 144, 160, 176, 192, 208, 224, 240, 255])  # 16-bit
            inds = np.digitize(img, bins)
            max_value = inds.max() + 1
            matrix_coocurrence = greycomatrix(inds, [1], [ np.pi/2],levels=max_value, normed=False, symmetric=False)

            f1 = contrast_feature(matrix_coocurrence)
            f2 = homogeneity_feature(matrix_coocurrence)
            f3= energy_feature(matrix_coocurrence)
            f4 = correlation_feature(matrix_coocurrence)
            f5 = asm_feature(matrix_coocurrence)

            glcmF_vector = []
            glcmF_vector.append(f1[0][0])
            glcmF_vector.append(f2[0][0])
            glcmF_vector.append(f3[0][0])
            glcmF_vector.append(f4[0][0])
            glcmF_vector.append(f5[0][0])

            for i in range(5):
             sheet.write(ite,i,glcmF_vector[i])


    workbook.close()
    print("finished GLCM")
    return glcmF_vector


# GLCM properties
def contrast_feature(matrix_coocurrence):
	contrast = greycoprops(matrix_coocurrence, 'contrast')
	return  contrast

def dissimilarity_feature(matrix_coocurrence):
	dissimilarity = greycoprops(matrix_coocurrence, 'dissimilarity')
	return dissimilarity

def homogeneity_feature(matrix_coocurrence):
	homogeneity = greycoprops(matrix_coocurrence, 'homogeneity')
	return  homogeneity

def energy_feature(matrix_coocurrence):
	energy = greycoprops(matrix_coocurrence, 'energy')
	return  energy

def correlation_feature(matrix_coocurrence):
	correlation = greycoprops(matrix_coocurrence, 'correlation')
	return  correlation

def asm_feature(matrix_coocurrence):
	asm = greycoprops(matrix_coocurrence, 'ASM')
	return  asm

def featureVector(filename1, filename2):
    print("Creating GLCM+DCT Feature Extraction.")
    array1 = read_excel(filename1)

    array2 = read_excel(filename2)
    tagged_array = np.concatenate((array1, array2),axis=1)
    print("Feature extraction vector ready")
    return tagged_array

def LabelfeatureVector(filename1, filename2):
    print("Creating Label Feature Extraction.")
    array1 = read_excel(filename1)
    array2 = read_excel(filename2)
    tagged_array = np.concatenate((array1, array2),axis=1)
    print("Feature extraction vector have tagged")
    return tagged_array

def dct(list_of_images, filename):

    workbook = xlsxwriter.Workbook(filename)
    sheet = workbook.add_worksheet('DCT')

    if sheet == '\0':
        print('error in file creation')
    else:
        print('DCT file created')
        for ite in range(len(list_of_images)):
            img=list_of_images[ite]
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            array = cv2.resize(gray, (256, 256))
            imf = np.float32(array)
            discrete_coeff = cv2.dct(imf)
            fp = zig_zag(discrete_coeff)

            for i in range(20):
                sheet.write(ite, i, fp[i])

    workbook.close()
    print("finished DCT")
def dct(list_of_images, filename):

    workbook = xlsxwriter.Workbook(filename)
    sheet = workbook.add_worksheet('DCT')

    if sheet == '\0':
        print('error in file creation')
    else:
        print('DCT file created')
        for ite in range(len(list_of_images)):
            img=list_of_images[ite]
            img=preprocessing2(img)[3]
            array = cv2.resize(img, (256, 256))
            imf = np.float32(array)
            discrete_coeff = cv2.dct(imf)
            fp = zig_zag(discrete_coeff)

            for i in range(20):
                sheet.write(ite, i, fp[i])

    workbook.close()
    print("finished DCT")




def load_images_from_folder(folder):
    print(" Database...")

    subfolder = os.listdir(folder)
    images = []
    label=[]

    for val in subfolder:

        path = os.path.join(folder, val)

        for image in os.listdir(path):
            if image.endswith(".jpg"):

               img = cv2.imread(os.path.join(path, image), 1)
               iclass_id=val
               label.append(iclass_id)
               viewGraphFourier(img)
               #viewGlcmPreprocessing(img)
               viewTemp(img)

               if img is not None:
                    images.append(img)

    print(len(images), " images have been read...")
    return images,label

def Dataframe_to_Excel(array, filename3):
    df = pd.DataFrame(array)
    writer = pd.ExcelWriter(filename3, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', header=False, index=False)
    writer.close()

def labell(val,filename):

    workbook = xlsxwriter.Workbook(filename)
    sheet = workbook.add_worksheet('label')

    if sheet == '\0':
     print('error in file creation')
    else:
     print('label file created')

     for i in range(len(val)-1):
         sheet.write(i, 0, label[i])
    workbook.close()


image_list,label= load_images_from_folder('Fabric')
image_list=(image_list)


filename1 = "Glcm.xlsx"
filename2 = "Dct.xlsx"
filename3 = "Feature.xlsx"
filename4=  'label.xlsx'
filename5='labelFeatureVector.xlsx'
filename6 = "Glcm_test.xlsx"
filename7 = "Dct_test.xlsx"
filename8 = "Feature_test.xlsx"
filename9='label_test.xlsx'
filename10='labelFeatureVector_test.xlsx'
filename11='Fastfourier.xlsx'

print('-'*15,'TRAIN','-'*15)
Glcm(image_list,filename1)
dct(image_list,filename2)
feature = featureVector(filename1, filename2)
Dataframe_to_Excel(feature, filename3)
labell(label,filename4)
featureLabel=LabelfeatureVector(filename3,filename4)
Dataframe_to_Excel(featureLabel,filename5)








